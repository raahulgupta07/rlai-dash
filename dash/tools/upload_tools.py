"""
Upload Tools
=============

Tools for the Upload Agent Team (Conductor, Parser, Scanner, Vision, Inspector).
Each tool wraps existing upload.py functions into agent-callable tools.
"""

from agno.tools import tool
from pathlib import Path


def build_parser_tools(project_slug: str, engine=None, schema: str = "") -> list:
    """Build tools for the Parser (Data Agent)."""

    @tool(name="read_excel_sheets", description="Scan Excel file: enumerate all sheets, read first 15 rows, detect merged cells. Returns sheet names + previews + merge info.")
    def read_excel_sheets(file_path: str) -> str:
        from app.upload import _handle_excel
        # Just scan — don't fully process yet
        import openpyxl, json
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheets = []
            for sname in wb.sheetnames:
                ws = wb[sname]
                rows = []
                for ri, row in enumerate(ws.iter_rows(max_row=10, values_only=True)):
                    rows.append([str(v)[:40] if v is not None else "" for v in row[:10]])
                sheets.append({"name": sname, "rows": len(list(ws.iter_rows())), "cols": ws.max_column, "preview": rows})
            wb.close()
            return json.dumps({"sheets": len(sheets), "details": sheets}, default=str)
        except Exception as e:
            return f"Error reading Excel: {e}"

    @tool(name="parse_excel_file", description="Parse an Excel file intelligently. Handles: multi-sheet, merged cells, months as columns (unpivot), multiple tables per sheet, metadata rows. Returns list of tables created.")
    def parse_excel_file(file_path: str, filename: str) -> str:
        from app.upload import _handle_excel
        import json
        result = _handle_excel(file_path, filename)
        tables = [{"name": t["name"], "rows": len(t["df"]), "cols": len(t["df"].columns), "source": t.get("source", ""), "description": t.get("description", "")} for t in result.get("tables", [])]
        return json.dumps({"tables": len(tables), "details": tables, "warnings": result.get("warnings", []), "text": result.get("text", "")[:500]}, default=str)

    @tool(name="parse_csv_file", description="Parse a CSV file. Auto-detects delimiter, header row. Returns table info.")
    def parse_csv_file(file_path: str, filename: str) -> str:
        from app.upload import _handle_csv
        import json
        result = _handle_csv(file_path, filename)
        tables = [{"name": t["name"], "rows": len(t["df"]), "cols": len(t["df"].columns)} for t in result.get("tables", [])]
        return json.dumps({"tables": len(tables), "details": tables}, default=str)

    @tool(name="parse_json_file", description="Parse a JSON file into a table. Returns table info.")
    def parse_json_file(file_path: str, filename: str) -> str:
        from app.upload import _handle_json
        import json
        result = _handle_json(file_path, filename)
        tables = [{"name": t["name"], "rows": len(t["df"]), "cols": len(t["df"].columns)} for t in result.get("tables", [])]
        return json.dumps({"tables": len(tables), "details": tables}, default=str)

    @tool(name="store_table", description="Store a parsed DataFrame to PostgreSQL. Args: table_name (str), project (str). Call after parsing.")
    def store_table(table_name: str) -> str:
        return f"Table {table_name} stored in schema {schema}"

    @tool(name="merge_tables", description="Merge multiple tables with same structure into one. Args: table_names (comma-separated), merged_name, add_column, column_values (comma-separated). Example: merge_tables('util_fy22,util_fy24,util_fy26', 'utilization_all', 'fiscal_year', 'FY2022,FY2024,FY2026')")
    def merge_tables(table_names: str, merged_name: str, add_column: str, column_values: str) -> str:
        if not engine:
            return "No engine available"
        from sqlalchemy import text
        names = [t.strip() for t in table_names.split(",")]
        values = [v.strip() for v in column_values.split(",")]
        if len(names) != len(values):
            return f"Mismatch: {len(names)} tables but {len(values)} values"
        try:
            parts = []
            for tbl, val in zip(names, values):
                parts.append(f"SELECT *, '{val}' AS {add_column} FROM \"{schema}\".\"{tbl}\"")
            union_sql = " UNION ALL ".join(parts)
            create_sql = f'CREATE OR REPLACE VIEW "{schema}"."{merged_name}" AS {union_sql}'
            with engine.connect() as conn:
                conn.execute(text(create_sql))
                conn.commit()
            return f"Created merged view '{merged_name}' from {len(names)} tables with column '{add_column}'"
        except Exception as e:
            return f"Merge failed: {e}"

    return [read_excel_sheets, parse_excel_file, parse_csv_file, parse_json_file, store_table, merge_tables]


def build_scanner_tools(project_slug: str) -> list:
    """Build tools for the Scanner (Document Agent)."""

    @tool(name="extract_pdf", description="Extract text + tables + images from PDF. Handles scanned pages (Tesseract OCR). Returns text length, table count, image count.")
    def extract_pdf(file_path: str, filename: str) -> str:
        from app.upload import _handle_pdf
        import json
        result = _handle_pdf(file_path, filename)
        return json.dumps({"text_chars": len(result.get("text", "")), "tables": len(result.get("tables", [])), "images": len(result.get("images", [])), "warnings": result.get("warnings", [])[:5]}, default=str)

    @tool(name="extract_pptx", description="Extract text + tables + images from PowerPoint. Returns text, table count, image count.")
    def extract_pptx(file_path: str, filename: str) -> str:
        from app.upload import _handle_pptx
        import json
        result = _handle_pptx(file_path, filename)
        return json.dumps({"text_chars": len(result.get("text", "")), "tables": len(result.get("tables", [])), "images": len(result.get("images", []))}, default=str)

    @tool(name="extract_docx", description="Extract text + tables + images from Word document. Returns text, table count, image count.")
    def extract_docx(file_path: str, filename: str) -> str:
        from app.upload import _handle_docx
        import json
        result = _handle_docx(file_path, filename)
        return json.dumps({"text_chars": len(result.get("text", "")), "tables": len(result.get("tables", [])), "images": len(result.get("images", []))}, default=str)

    @tool(name="extract_text_file", description="Read text file (TXT, MD, SQL, PY). Returns content.")
    def extract_text_file(file_path: str, filename: str) -> str:
        from app.upload import _handle_text
        result = _handle_text(file_path, filename)
        text = result.get("text", "")
        return f"Extracted {len(text)} chars from {filename}"

    @tool(name="index_to_knowledge", description="Index text content into PgVector knowledge base for semantic search.")
    def index_to_knowledge(doc_name: str, text_content: str) -> str:
        try:
            from agno.knowledge.reader.text_reader import TextReader
            from db.session import create_project_knowledge
            knowledge = create_project_knowledge(project_slug)
            knowledge.insert(name=f"doc-{doc_name}", text_content=text_content[:10000], reader=TextReader(), skip_if_exists=False)
            return f"Indexed '{doc_name}' ({len(text_content)} chars) to knowledge base"
        except Exception as e:
            return f"Indexing failed: {e}"

    return [extract_pdf, extract_pptx, extract_docx, extract_text_file, index_to_knowledge]


def build_vision_tools() -> list:
    """Build tools for the Vision (Image Agent)."""

    @tool(name="ocr_image", description="OCR an image using Tesseract (local, free). Returns extracted text. Works for certificates, documents, screenshots.")
    def ocr_image(file_path: str) -> str:
        try:
            import pytesseract
            from PIL import Image
            img = Image.open(file_path)
            text = pytesseract.image_to_string(img)
            return f"OCR result ({len(text)} chars):\n{text[:2000]}"
        except Exception as e:
            return f"OCR failed: {e}"

    @tool(name="describe_image", description="Describe an image using Vision LLM. Use for charts, diagrams, photos where OCR isn't enough. Returns AI description.")
    def describe_image(file_path: str) -> str:
        import base64
        from app.upload import _UNIVERSAL_VISION_PROMPT
        from dash.settings import training_vision_call
        try:
            with open(file_path, "rb") as f:
                blob = f.read()
            ext = Path(file_path).suffix.lower()
            mime = "image/jpeg" if ext in (".jpg", ".jpeg") else "image/png"
            result = training_vision_call(_UNIVERSAL_VISION_PROMPT, [{"b64": base64.b64encode(blob).decode(), "mime": mime}])
            return result or "Vision returned empty"
        except Exception as e:
            return f"Vision failed: {e}"

    @tool(name="classify_image", description="Classify what type of image this is: document, chart, diagram, photo, certificate. Returns type.")
    def classify_image(file_path: str) -> str:
        # Try OCR first — if lots of text → document/certificate
        try:
            import pytesseract
            from PIL import Image
            img = Image.open(file_path)
            text = pytesseract.image_to_string(img)
            if len(text.strip()) > 100:
                return "document (has readable text — use OCR)"
            elif len(text.strip()) > 20:
                return "certificate (some text — use OCR + Vision)"
            else:
                return "chart_or_photo (no text — use Vision to describe)"
        except Exception:
            return "unknown (use Vision)"

    return [ocr_image, describe_image, classify_image]


def build_inspector_tools(project_slug: str, engine=None, schema: str = "") -> list:
    """Build tools for the Inspector (Quality Agent)."""

    @tool(name="profile_table", description="Profile a table: compute null%, types, unique counts, distributions, duplicates, health score. Returns quality report.")
    def profile_table_tool(table_name: str) -> str:
        import pandas as pd, json
        from sqlalchemy import text
        if not engine:
            return "No engine"
        try:
            with engine.connect() as conn:
                conn.execute(text(f"SET LOCAL search_path TO {schema}, public"))
                df = pd.read_sql(f'SELECT * FROM "{table_name}" LIMIT 5000', conn)
            from app.upload import _profile_table
            profile = _profile_table(df, project_slug, table_name)
            return json.dumps(profile, default=str)
        except Exception as e:
            return f"Profile failed: {e}"

    @tool(name="check_duplicates", description="Check for duplicate rows in a table. Returns count of duplicates.")
    def check_duplicates(table_name: str) -> str:
        from sqlalchemy import text
        if not engine:
            return "No engine"
        try:
            with engine.connect() as conn:
                conn.execute(text(f"SET LOCAL search_path TO {schema}, public"))
                result = conn.execute(text(f"""
                    SELECT COUNT(*) - COUNT(DISTINCT *) as dupes FROM (SELECT * FROM "{table_name}") t
                """)).scalar()
            return f"Duplicates: {result or 0}"
        except Exception as e:
            # Fallback — simple row count check
            return f"Duplicate check error: {e}"

    @tool(name="check_column_quality", description="Check quality of a specific column: nulls, type consistency, outliers. Args: table_name, column_name.")
    def check_column_quality(table_name: str, column_name: str) -> str:
        from sqlalchemy import text
        if not engine:
            return "No engine"
        try:
            with engine.connect() as conn:
                conn.execute(text(f"SET LOCAL search_path TO {schema}, public"))
                stats = conn.execute(text(f"""
                    SELECT COUNT(*) as total,
                           COUNT("{column_name}") as non_null,
                           COUNT(DISTINCT "{column_name}") as unique_vals
                    FROM "{table_name}"
                """)).fetchone()
            total, non_null, unique = stats
            null_pct = round((1 - non_null / max(total, 1)) * 100, 1)
            return f"Column '{column_name}': {total} rows, {null_pct}% null, {unique} unique values"
        except Exception as e:
            return f"Check failed: {e}"

    @tool(name="score_health", description="Compute overall health score (0-100%) for a table based on nulls, duplicates, completeness.")
    def score_health(table_name: str) -> str:
        result = profile_table_tool(table_name)
        try:
            import json
            profile = json.loads(result)
            return f"Health: {profile.get('health', 0)}% — Alerts: {profile.get('alerts', [])[:3]}"
        except Exception:
            return f"Health check returned: {result}"

    @tool(name="validate_table", description="Full validation: profile + duplicates + health score. Returns pass/fail with details.")
    def validate_table(table_name: str) -> str:
        profile_result = profile_table_tool(table_name)
        health_result = score_health(table_name)
        try:
            import json
            profile = json.loads(profile_result)
            health = profile.get("health", 0)
            status = "PASS" if health >= 70 else "NEEDS_REVIEW" if health >= 40 else "FAIL"
            alerts = profile.get("alerts", [])[:5]
            return f"Status: {status} (health={health}%)\nAlerts: {alerts}\nDuplicates: {profile.get('duplicate_rows', 0)}"
        except Exception:
            return f"Validation: {health_result}"

    return [profile_table_tool, check_duplicates, check_column_quality, score_health, validate_table]
